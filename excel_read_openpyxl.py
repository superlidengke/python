import openpyxl

file = r"C:\Users\UC212310\Desktop\work\ESG\18592_2017_20180423_030645.xlsx"
book = openpyxl.load_workbook(file, read_only=True)
# current active sheet
sheet = book.active
cell_a1 = sheet['A1']
print(cell_a1.value)
cell_a1 = sheet.cell(row=1,column=1)
print(cell_a1.value)

# iterate row
for row in sheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value,end=" ")
    print()
    break

# get sheet by name
sheet = book["Sheet"]
row = list(sheet.rows)[1]
print(row[1].value)